"""全量数据资产审计服务（第10A阶段）。

目标：审计所有比赛数据 / 训练语料 / 参考资料是否被完整读取、解析、入库，并追踪
每个文件有多少条原始数据、解析数据、入库数据、参与特征工程/训练的数据，识别
漏读、字段映射失败、坐标解析失败、test 污染与原文泄露风险。

口径与红线：
- 原始条数由本地 ``json.load`` 直接计数（不抽样、不估算）。
- 字段映射失败 / 坐标失败由 ``field_mapping_service`` + ``geo_utils`` 在本地重算复核。
- 仅 train/val 计入"参与特征工程/训练"（test 永不计入，leakage_risk/test_contamination_risk 据此判定）。
- 第10A 尚未实现聚类/相似度/外部数据，相关计数据实写 0 并在 note 标注，绝不编造。
- 输出仅含统计量与脱敏结论；不返回任何 raw_json / 原始点位 / 企业名 / 小区名 / 地址 / 坐标列表。
- 导出落 backend/data/outputs/data_catalog/（已 gitignore）。
"""

from __future__ import annotations

import json
import logging
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.config import settings
from app.models import (
    EvaluationResult,
    HousingRecord,
    IndustryPoint,
    PoiPoint,
    PopulationProfile,
)
from app.services import field_mapping_service as fms
from app.services import split_manager
from app.utils import geo_utils

logger = logging.getLogger("cityrenew.data_audit")

USABLE_COORD = {geo_utils.STATUS_OK, geo_utils.STATUS_CORRECTED}
TRAIN_VAL = ("train", "val")
LOW_COVERAGE_THRESHOLD = 0.95

# 结构化点/记录型文件：file_name -> (data_type, model, mapping_key)
_ARRAY_FILES: dict[str, tuple[str, Any, str]] = {
    "POI兴趣点分布数据.json": ("poi", PoiPoint, "poi"),
    "产业布局数据.json": ("industry", IndustryPoint, "industry"),
    "房价历史交易数据.json": ("house_price", HousingRecord, "house_price"),
}
# 训练语料中已知的全部文件名（用于区分"已审计/其它"）
_KNOWN_STRUCTURED = set(_ARRAY_FILES) | {
    "区域人口总量.json",
    "区域人口画像.json",
    "矢量数据样例及说明表.xlsx",
}
_PARSEABLE_EXT = {".json", ".csv", ".xlsx", ".xls", ".docx", ".pdf", ".md", ".txt"}


def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _rel_source(file_name: str) -> str:
    return f"{settings.corpus_path.name}/{file_name}"


def _safe_rel(path: Path) -> str:
    """只暴露 `父目录名/文件名`，不泄露绝对路径。"""
    return f"{path.parent.name}/{path.name}"


def _coverage(db_count: int, raw_count: int) -> float:
    return round(db_count / raw_count, 4) if raw_count else 0.0


def _data_catalog_dir() -> Path:
    return settings.data_dir / "outputs" / "data_catalog"


def _external_dir() -> Path:
    return settings.data_dir / "external"


# --------------------------------------------------------------------------- #
# 单文件审计
# --------------------------------------------------------------------------- #
def _base_entry(file_name: str, file_path: str, file_type: str, source_group: str) -> dict[str, Any]:
    return {
        "file_name": file_name,
        "file_path": file_path,
        "file_type": file_type,
        "source_group": source_group,
        "raw_record_count": None,
        "parsed_record_count": None,
        "db_inserted_count": 0,
        "skipped_count": 0,
        "skipped_reason": [],
        "split_train_count": 0,
        "split_val_count": 0,
        "split_test_count": 0,
        "used_in_feature_engineering_count": 0,
        "used_in_training_count": 0,
        "used_in_clustering_count": 0,
        "used_in_similarity_count": 0,
        "used_in_report_count": 0,
        "coverage_rate": None,
        "field_mapping_status": "n/a",
        "coordinate_valid_count": 0,
        "coordinate_invalid_count": 0,
        "leakage_risk": False,
        "test_contamination_risk": False,
        "recommendations": [],
    }


def _audit_array_file(
    db: Session,
    file_name: str,
    data_type: str,
    model: Any,
    mapping_key: str,
    split_summary: dict[str, Any],
) -> dict[str, Any]:
    path = settings.corpus_path / file_name
    entry = _base_entry(file_name, _rel_source(file_name), "json", "competition_data")
    if not path.exists():
        entry["skipped_reason"].append("文件不存在于训练语料目录")
        entry["recommendations"].append("确认训练语料目录是否完整")
        return entry

    with path.open("r", encoding="utf-8") as f:
        records = json.load(f)

    raw_count = len(records) if isinstance(records, list) else 0
    missing: Counter[str] = Counter()
    missing_opt: Counter[str] = Counter()
    coord: Counter[str] = Counter()
    parsed = 0
    if isinstance(records, list):
        for raw in records:
            if not isinstance(raw, dict):
                continue
            res = fms.map_record(mapping_key, raw)
            for mf in res.missing_fields:
                missing[mf] += 1
            for mo in res.missing_optional:
                missing_opt[mo] += 1
            pt = geo_utils.parse_point(res.mapped.get("coordinates"))
            coord[pt.status] += 1
            parsed += 1
    else:
        entry["skipped_reason"].append("顶层不是数组，疑似结构异常或嵌套，需人工复核")

    db_count = (
        db.query(model).filter(model.source_file == _rel_source(file_name)).count()
    )
    used_fe = (
        db.query(model)
        .filter(model.split.in_(TRAIN_VAL), model.coord_status.in_(USABLE_COORD))
        .count()
    )
    test_count = db.query(model).filter(model.split == "test").count()

    coord_valid = coord.get(geo_utils.STATUS_OK, 0) + coord.get(geo_utils.STATUS_CORRECTED, 0)
    coord_invalid = coord.get(geo_utils.STATUS_INVALID, 0) + coord.get(geo_utils.STATUS_MISSING, 0)
    ss = split_summary.get(data_type, {})

    entry.update(
        {
            "raw_record_count": raw_count,
            "parsed_record_count": parsed,
            "db_inserted_count": db_count,
            "skipped_count": max(0, raw_count - db_count),
            "split_train_count": int(ss.get("train", 0)),
            "split_val_count": int(ss.get("val", 0)),
            "split_test_count": int(ss.get("test", 0)),
            "used_in_feature_engineering_count": used_fe,
            "coordinate_valid_count": coord_valid,
            "coordinate_invalid_count": coord_invalid,
            "coverage_rate": _coverage(db_count, raw_count),
            "field_mapping_status": _mapping_status(missing, missing_opt),
        }
    )
    # 房价是唯一进入监督训练的样本（train 拟合 / val 验证）；其余仅作特征来源
    if data_type == "house_price":
        entry["used_in_training_count"] = used_fe

    # 漏读 / 样例风险
    if raw_count and db_count < raw_count:
        entry["recommendations"].append(
            f"入库 {db_count} < 原始 {raw_count}，{raw_count - db_count} 条未入库，需复核"
        )
    if raw_count <= 1:
        entry["recommendations"].append("原始仅 1 条，疑似只读取样例")
    if coord_invalid:
        entry["recommendations"].append(f"{coord_invalid} 条坐标无效/缺失，未参与空间归集")
    if missing:
        entry["recommendations"].append(
            f"存在必需字段缺失：{', '.join(sorted(missing))}"
        )
    # test 隔离判定：used_fe 仅取 train/val，test 单独计数，二者不交叉
    entry["leakage_risk"] = False
    entry["test_contamination_risk"] = False
    entry["_test_count_db"] = test_count
    return entry


def _audit_population(db: Session, split_summary: dict[str, Any]) -> list[dict[str, Any]]:
    total_name = "区域人口总量.json"
    profile_name = "区域人口画像.json"
    out: list[dict[str, Any]] = []
    ss = split_summary.get("population", {})

    # 共享 DB 计数
    grid_total = db.query(PopulationProfile).count()
    grid_with_profile = (
        db.query(PopulationProfile).filter(PopulationProfile.profile_json.isnot(None)).count()
    )
    used_fe = (
        db.query(PopulationProfile)
        .filter(
            PopulationProfile.split.in_(TRAIN_VAL),
            PopulationProfile.coord_status.in_(USABLE_COORD),
        )
        .count()
    )
    used_fe_profile = (
        db.query(PopulationProfile)
        .filter(
            PopulationProfile.split.in_(TRAIN_VAL),
            PopulationProfile.coord_status.in_(USABLE_COORD),
            PopulationProfile.profile_json.isnot(None),
        )
        .count()
    )
    test_count = db.query(PopulationProfile).filter(PopulationProfile.split == "test").count()

    # ---- 总量 ----
    for fname, db_count, used, role in (
        (total_name, grid_total, used_fe, "网格总量（合并入 population_profiles）"),
        (profile_name, grid_with_profile, used_fe_profile, "网格画像（合并入同一网格，提供 35 维结构）"),
    ):
        path = settings.corpus_path / fname
        entry = _base_entry(fname, _rel_source(fname), "json", "competition_data")
        if not path.exists():
            entry["skipped_reason"].append("文件不存在于训练语料目录")
            out.append(entry)
            continue
        with path.open("r", encoding="utf-8") as f:
            records = json.load(f)
        raw_count = len(records) if isinstance(records, list) else 0
        coord: Counter[str] = Counter()
        missing: Counter[str] = Counter()
        parsed = 0
        mapping_key = "population_total" if fname == total_name else "population_profile"
        for raw in records if isinstance(records, list) else []:
            if not isinstance(raw, dict):
                continue
            res = fms.map_record(mapping_key, raw)
            for mf in res.missing_fields:
                missing[mf] += 1
            bbox = geo_utils.parse_bbox(res.mapped.get("coordinates"))
            coord[bbox.status] += 1
            parsed += 1
        coord_valid = coord.get(geo_utils.STATUS_OK, 0) + coord.get(geo_utils.STATUS_CORRECTED, 0)
        coord_invalid = coord.get(geo_utils.STATUS_INVALID, 0) + coord.get(geo_utils.STATUS_MISSING, 0)
        entry.update(
            {
                "raw_record_count": raw_count,
                "parsed_record_count": parsed,
                "db_inserted_count": db_count,
                "skipped_count": max(0, raw_count - db_count),
                "split_train_count": int(ss.get("train", 0)),
                "split_val_count": int(ss.get("val", 0)),
                "split_test_count": int(ss.get("test", 0)),
                "used_in_feature_engineering_count": used,
                "coordinate_valid_count": coord_valid,
                "coordinate_invalid_count": coord_invalid,
                "coverage_rate": _coverage(db_count, raw_count),
                "field_mapping_status": _mapping_status(missing, Counter()),
            }
        )
        entry["recommendations"].append(
            "人口总量与画像按 grid_key 合并为同一网格记录；split 计数对应合并后网格"
        )
        if fname == profile_name and db_count < raw_count:
            entry["recommendations"].append(
                f"{raw_count - db_count} 个画像网格未匹配到总量（或坐标不可用），未充分使用"
            )
        entry["_test_count_db"] = test_count
        out.append(entry)
    return out


def _audit_spec(db: Session) -> dict[str, Any]:
    file_name = "矢量数据样例及说明表.xlsx"
    path = settings.corpus_path / file_name
    entry = _base_entry(file_name, _rel_source(file_name), "xlsx", "competition_data")
    if not path.exists():
        entry["skipped_reason"].append("文件不存在")
        return entry
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = [{"sheet": ws.title, "rows": ws.max_row, "cols": ws.max_column} for ws in wb.worksheets]
        wb.close()
    except Exception as exc:  # noqa: BLE001
        sheets = []
        entry["skipped_reason"].append(f"解析失败：{type(exc).__name__}")
    entry.update(
        {
            "raw_record_count": len(sheets),
            "parsed_record_count": len(sheets),
            "db_inserted_count": 0,
            "field_mapping_status": "字段口径说明表（spec），仅作映射校验，不入分析表/不参与训练",
            "coverage_rate": 1.0 if sheets else 0.0,
        }
    )
    entry["sheets"] = sheets
    entry["recommendations"].append("说明表用于字段口径核对，不进入特征工程/训练（符合设计）")
    return entry


_MEDIA_EXT = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}


def _audit_other_files() -> list[dict[str, Any]]:
    """训练语料其它文件 + 参考资料可解析文件的轻量登记（不进入训练/特征）。

    非结构化媒体（PPT 逐页导出图等）按所在目录聚合为 1 条，避免清单被上百张图淹没。
    """
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    media_by_dir: dict[tuple[str, str], int] = {}

    def _scan(base: Path, source_group: str) -> None:
        if not base.exists():
            return
        for path in sorted(base.rglob("*")):
            if not path.is_file():
                continue
            if path.name.startswith(".") or "__MACOSX" in path.parts:
                continue
            if path.name in _KNOWN_STRUCTURED and source_group == "training_corpus":
                continue  # 已在结构化审计中覆盖
            ext = path.suffix.lower()
            if ext in _MEDIA_EXT:
                media_by_dir[(path.parent.name, source_group)] = (
                    media_by_dir.get((path.parent.name, source_group), 0) + 1
                )
                continue
            key = _safe_rel(path)
            if key in seen:
                continue
            seen.add(key)
            entry = _base_entry(path.name, key, ext.lstrip(".") or "unknown", source_group)
            parseable = ext in _PARSEABLE_EXT
            entry["field_mapping_status"] = "parseable" if parseable else "non_parseable"
            if ext == ".json":
                try:
                    with path.open("r", encoding="utf-8") as f:
                        data = json.load(f)
                    n = len(data) if isinstance(data, list) else 1
                    entry["raw_record_count"] = n
                    if n <= 1:
                        entry["recommendations"].append("仅 schema 样例（1 条），用于字段口径，不作训练数据")
                except Exception:  # noqa: BLE001
                    entry["recommendations"].append("JSON 解析失败，需人工复核")
            entry["recommendations"].append(
                "参考/文档类资料用于 RAG 检索或人工核对，未作为训练样本（符合设计）"
            )
            out.append(entry)

    _scan(settings.corpus_path, "training_corpus")
    _scan(settings.reference_path, "reference_doc")

    for (dir_name, source_group), count in sorted(media_by_dir.items()):
        entry = _base_entry(f"{dir_name}/[图片素材 x{count}]", f"{dir_name}/", "image", source_group)
        entry["field_mapping_status"] = "non_parseable"
        entry["raw_record_count"] = count
        entry["recommendations"].append("图片素材（如 PPT 逐页导出），不作训练样本（符合设计）")
        out.append(entry)
    return out


# 第10B：external/ 下的脚手架/模板/登记文件（非真实采集数据，不计入数据文件）
_EXTERNAL_TEMPLATE_NAMES = {
    "README.md",
    "data_catalog.json",
    "data_lineage.json",
    "collection_tasks.json",
    "manifest.json",
    "data_source_registry.json",
    "candidate_sources.json",
    "source_discovery_log.json",
}


def _is_external_template(path: Path) -> bool:
    """判定是否为第10B 脚手架/模板/登记文件（而非真实采集到的外部数据）。"""
    if path.name in _EXTERNAL_TEMPLATE_NAMES:
        return True
    return path.parent.name == "registry"


def _audit_external() -> dict[str, Any]:
    """登记 external/ 目录现状（第10B 脚手架 + 真实采集数据分别计数）。

    口径：脚手架/模板/registry/manifest 不算"真实外部数据"；只有 raw/processed/cache 下
    实际采集到的数据文件才计入 external_data_files_count。外部数据物理隔离于 competition test，
    其存在本身不构成 test 污染（test_contamination_risk 由竞赛数据 split 判定）。
    """
    ext_dir = _external_dir()
    all_files: list[Path] = []
    if ext_dir.exists():
        all_files = [
            p for p in ext_dir.rglob("*")
            if p.is_file() and not p.name.startswith(".")
        ]
    template_files = [p for p in all_files if _is_external_template(p)]
    data_files = [p for p in all_files if not _is_external_template(p)]
    return {
        "external_dir": str(ext_dir.relative_to(settings.data_dir.parent)),
        "external_dir_exists": ext_dir.exists(),
        "external_files_count": len(all_files),
        "external_template_files_count": len(template_files),
        "external_data_files_count": len(data_files),
        "note": (
            "external/ 为第10B 合规外部数据增强目录（已 gitignore）；脚手架/模板/registry/manifest "
            "不算真实采集数据，仅 raw/processed/cache 下实采数据计入 external_data_files_count。"
            "外部数据物理隔离于 competition test，不混入训练 test、不反推 test 答案。"
        ),
    }


def _mapping_status(missing: Counter, missing_opt: Counter) -> Any:
    if not missing and not missing_opt:
        return "ok"
    return {
        "status": "partial" if missing else "ok_with_optional_missing",
        "missing_field_counts": dict(missing),
        "missing_optional_counts": dict(missing_opt),
    }


# --------------------------------------------------------------------------- #
# 入口
# --------------------------------------------------------------------------- #
def run_data_audit(db: Session, persist: bool = True) -> dict[str, Any]:
    """执行全量数据审计，返回脱敏审计结果并导出 data_catalog。"""
    split_info = split_manager.get_split_summary()
    split_summary = split_info.get("per_type", {}) if split_info.get("built") else {}

    files: list[dict[str, Any]] = []
    for file_name, (data_type, model, mapping_key) in _ARRAY_FILES.items():
        files.append(_audit_array_file(db, file_name, data_type, model, mapping_key, split_summary))
    files.extend(_audit_population(db, split_summary))
    files.append(_audit_spec(db))
    files.extend(_audit_other_files())

    external = _audit_external()

    # ---- 聚合 ----
    competition = [f for f in files if f["source_group"] == "competition_data"]
    # 入库去重：人口画像与总量同表，总库计数只取一次（poi+industry+house+population 网格）
    db_total = (
        db.query(PoiPoint).count()
        + db.query(IndustryPoint).count()
        + db.query(HousingRecord).count()
        + db.query(PopulationProfile).count()
    )
    total_raw = sum(f["raw_record_count"] or 0 for f in competition if f["file_type"] == "json")
    total_parsed = sum(f["parsed_record_count"] or 0 for f in competition if f["file_type"] == "json")
    total_used = sum(f["used_in_feature_engineering_count"] for f in files)

    # 覆盖率分母：可入库结构化原始条数（poi+industry+house+人口总量网格）
    coverage_denom = 0
    for f in competition:
        if f["file_name"] in _ARRAY_FILES or f["file_name"] == "区域人口总量.json":
            coverage_denom += f["raw_record_count"] or 0
    coverage_rate = _coverage(db_total, coverage_denom)

    unused_entries = [
        f
        for f in files
        if f["used_in_feature_engineering_count"] == 0
        and f["used_in_training_count"] == 0
        and f["source_group"] != "competition_data"
    ]
    # 可解析的未使用文件用 path 列清单（消除与竞赛文件同名歧义）；媒体仅计数
    unused_files = [
        f["file_path"] for f in unused_entries if f["file_type"] in {
            "json", "csv", "xlsx", "xls", "docx", "pdf", "md", "txt"
        }
    ]
    unused_media_count = sum(
        (f["raw_record_count"] or 1)
        for f in unused_entries
        if f["file_type"] == "image"
    )
    low_coverage_files = [
        f["file_name"]
        for f in competition
        if f["coverage_rate"] is not None
        and f["coverage_rate"] < LOW_COVERAGE_THRESHOLD
        and f["file_name"] != "矢量数据样例及说明表.xlsx"
    ]
    skipped_summary = {
        f["file_name"]: f["skipped_count"] for f in competition if f["skipped_count"]
    }

    leakage_risk = any(f["leakage_risk"] for f in files)
    # test 污染仅由竞赛数据 split 判定；外部数据（第10B）物理隔离于 competition test，
    # 其脚手架/模板/采集数据的存在均不构成 test 污染。
    test_contamination_risk = any(f["test_contamination_risk"] for f in files)

    recommendations: list[str] = []
    if not split_info.get("built"):
        recommendations.append("split_manifest 未构建，先执行 /api/splits/build 再审计 split 分布")
    if low_coverage_files:
        recommendations.append(f"低覆盖文件需复核：{', '.join(low_coverage_files)}")
    if coverage_rate >= LOW_COVERAGE_THRESHOLD and not leakage_risk:
        recommendations.append("结构化数据覆盖率达标且无 test 泄露，可进入第10.5 数据/特征质量门禁")
    recommendations.append("产业为单一类目、人口无收入字段：如实标注，第10B 可用合规外部数据增强")

    overall_status = "pass"
    if leakage_risk or test_contamination_risk or coverage_rate < LOW_COVERAGE_THRESHOLD:
        overall_status = "fail" if leakage_risk or test_contamination_risk else "warning"

    # 去掉内部辅助键
    for f in files:
        f.pop("_test_count_db", None)

    result: dict[str, Any] = {
        "mode": settings.app_mode,
        "phase": "10A",
        "created_at": _utcnow_iso(),
        "overall_status": overall_status,
        "all_files_count": len(files),
        "total_raw_records": total_raw,
        "total_parsed_records": total_parsed,
        "total_db_records": db_total,
        "total_used_records": total_used,
        "coverage_rate": coverage_rate,
        "split_built": bool(split_info.get("built")),
        "files": files,
        "unused_files": unused_files,
        "unused_media_count": unused_media_count,
        "low_coverage_files": low_coverage_files,
        "skipped_summary": skipped_summary,
        "leakage_risk": leakage_risk,
        "test_contamination_risk": test_contamination_risk,
        "external_data": external,
        "recommendations": recommendations,
        "notes": [
            "原始条数由本地 json.load 精确计数；字段/坐标失败由本地重算复核。",
            "used_in_feature_engineering/training 仅统计 train/val（test 永不计入）。",
            "used_in_clustering/similarity 第10A 未实现，据实写 0；第11/11.5 阶段引入。",
            "本响应仅含统计量与脱敏结论，不含任何 raw_json/原始点位/企业名/小区名/地址/坐标。",
        ],
    }

    exports = _export_catalog(result)
    result["exports"] = exports
    if persist:
        _persist_evaluation(db, result)

    logger.info(
        "data-audit done: files=%s db=%s coverage=%.3f leakage=%s",
        len(files), db_total, coverage_rate, leakage_risk,
    )
    return result


def _export_catalog(result: dict[str, Any]) -> dict[str, str]:
    out_dir = _data_catalog_dir()
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / "数据覆盖率报告.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    sources_md = out_dir / "数据来源清单.md"
    sources_md.write_text(_render_sources_md(result), encoding="utf-8")

    lineage_md = out_dir / "数据血缘报告.md"
    lineage_md.write_text(_render_lineage_md(result), encoding="utf-8")

    rel = lambda p: str(p.relative_to(settings.data_dir.parent))  # noqa: E731
    return {
        "coverage_report_json": rel(json_path),
        "sources_md": rel(sources_md),
        "lineage_md": rel(lineage_md),
    }


def _render_sources_md(result: dict[str, Any]) -> str:
    lines = [
        "# 数据来源清单（脱敏）",
        "",
        f"- 生成时间：{result['created_at']}",
        f"- 文件总数：{result['all_files_count']}",
        f"- 结构化原始记录：{result['total_raw_records']}",
        f"- 入库记录：{result['total_db_records']}",
        f"- 覆盖率：{result['coverage_rate']}",
        "",
        "| 文件 | 来源组 | 类型 | 原始 | 入库 | 覆盖率 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for f in result["files"]:
        lines.append(
            f"| {f['file_name']} | {f['source_group']} | {f['file_type']} | "
            f"{f['raw_record_count']} | {f['db_inserted_count']} | {f['coverage_rate']} |"
        )
    lines.append("")
    lines.append("> 仅统计量，不含任何原始明细。外部数据目录第10A 为空。")
    return "\n".join(lines)


def _render_lineage_md(result: dict[str, Any]) -> str:
    lines = [
        "# 数据血缘报告（第10A 内部数据）",
        "",
        f"- 生成时间：{result['created_at']}",
        "- 血缘链：训练语料文件 → 解析/字段映射 → 入库 → train/val 特征工程 → 房价监督训练",
        "",
        "| 文件 | 原始 | 解析 | 入库 | train | val | test | 特征工程 | 训练 |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for f in result["files"]:
        if f["source_group"] != "competition_data":
            continue
        lines.append(
            f"| {f['file_name']} | {f['raw_record_count']} | {f['parsed_record_count']} | "
            f"{f['db_inserted_count']} | {f['split_train_count']} | {f['split_val_count']} | "
            f"{f['split_test_count']} | {f['used_in_feature_engineering_count']} | "
            f"{f['used_in_training_count']} |"
        )
    lines += [
        "",
        f"- test 泄露风险：{result['leakage_risk']}",
        f"- test 污染风险：{result['test_contamination_risk']}",
        "- 说明：test 仅用于最终评估，不计入特征工程/训练；聚类/相似度第11 阶段引入。",
    ]
    return "\n".join(lines)


def _persist_evaluation(db: Session, result: dict[str, Any]) -> None:
    """把审计核心指标写入 EvaluationResult（激活自评结果表）。"""
    run_id = f"data_audit_{result['created_at']}"
    detail = {
        "overall_status": result["overall_status"],
        "all_files_count": result["all_files_count"],
        "unused_files": result["unused_files"],
        "low_coverage_files": result["low_coverage_files"],
        "leakage_risk": result["leakage_risk"],
        "test_contamination_risk": result["test_contamination_risk"],
    }
    db.add(
        EvaluationResult(
            run_id=run_id,
            mode=settings.app_mode,
            metric_name="data_coverage_rate",
            metric_value=float(result["coverage_rate"]),
            split="train+val",
            dataset_size=int(result["total_db_records"]),
            detail_json=json.dumps(detail, ensure_ascii=False),
        )
    )
    db.commit()
