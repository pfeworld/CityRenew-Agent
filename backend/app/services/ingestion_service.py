"""本地结构化资料导入服务（第2阶段）。

流程：扫描训练语料目录 → 识别 JSON/XLSX → 字段映射 + 坐标解析 → 写入对应表
→ 写入 DataFile → 生成数据质量报告（仅统计量，无原文）。

红线：
- 仅本地读取；不调用任何外部 API；不做 RAG / 四维分析 / 报告。
- 日志与返回只含：文件名、记录数、缺失字段统计、坐标状态计数、错误类型；不输出语料原文。
- 派生产物（quality_report.json）落 backend/data/processed/（已 gitignore）。
"""

from __future__ import annotations

import hashlib
import json
import logging
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.config import settings
from app.database import init_db
from app.models import (
    DataFile,
    HousingRecord,
    IndustryPoint,
    PoiPoint,
    PopulationProfile,
)
from app.services import field_mapping_service as fms
from app.utils import geo_utils

logger = logging.getLogger("cityrenew.ingestion")

# 文件名 → data_type（本阶段仅处理结构化数据；PPT/案例/政策不在此列）
KNOWN_FILES: dict[str, str] = {
    "POI兴趣点分布数据.json": "poi",
    "产业布局数据.json": "industry",
    "房价历史交易数据.json": "house_price",
    "区域人口总量.json": "population_total",
    "区域人口画像.json": "population_profile",
    "矢量数据样例及说明表.xlsx": "spec",
}

_POINT_MODEL = {"poi": PoiPoint, "industry": IndustryPoint}


def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def _rel_source(file_name: str) -> str:
    return f"{settings.corpus_path.name}/{file_name}"


def _int_or_none(value: Any) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _float_or_none(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------------------- #
# 清理（幂等）
# --------------------------------------------------------------------------- #
def _clear_tables(db: Session) -> None:
    for model in (PoiPoint, IndustryPoint, HousingRecord, PopulationProfile):
        db.query(model).delete()
    # 仅清理训练语料来源的 DataFile，保留其它（当前无其它）
    db.query(DataFile).filter(DataFile.source == "corpus").delete()
    db.flush()


# --------------------------------------------------------------------------- #
# 点数据（POI / 产业）
# --------------------------------------------------------------------------- #
def _ingest_point_file(db: Session, file_name: str, data_type: str) -> dict[str, Any]:
    path = settings.corpus_path / file_name
    model = _POINT_MODEL[data_type]
    with path.open("r", encoding="utf-8") as f:
        records: list[dict] = json.load(f)

    coord_stats: Counter[str] = Counter()
    missing_counts: Counter[str] = Counter()
    written = 0
    for idx, raw in enumerate(records):
        result = fms.map_record(data_type, raw)
        for mf in result.missing_fields:
            missing_counts[mf] += 1
        pt = geo_utils.parse_point(result.mapped.get("coordinates"))
        coord_stats[pt.status] += 1
        source_id = result.mapped.get("source_id") or f"{data_type}_{idx:06d}"
        db.add(
            model(
                source_id=str(source_id),
                name=result.mapped.get("name"),
                category_name=result.mapped.get("category_name"),
                district_name=result.mapped.get("district_name"),
                address=result.mapped.get("address"),
                lng=pt.lng if pt.is_usable else None,
                lat=pt.lat if pt.is_usable else None,
                coord_status=pt.status,
                source_file=_rel_source(file_name),
                raw_json=result.raw_json,
            )
        )
        written += 1

    db.flush()
    return _file_report(
        file_name, data_type, path, len(records), written,
        missing_counts, coord_stats, {},
    )


# --------------------------------------------------------------------------- #
# 房价
# --------------------------------------------------------------------------- #
def _ingest_house_file(db: Session, file_name: str) -> dict[str, Any]:
    path = settings.corpus_path / file_name
    with path.open("r", encoding="utf-8") as f:
        records: list[dict] = json.load(f)

    coord_stats: Counter[str] = Counter()
    missing_counts: Counter[str] = Counter()
    missing_optional_counts: Counter[str] = Counter()
    written = 0
    for idx, raw in enumerate(records):
        result = fms.map_record("house_price", raw)
        for mf in result.missing_fields:
            missing_counts[mf] += 1
        for mo in result.missing_optional:
            missing_optional_counts[mo] += 1
        pt = geo_utils.parse_point(result.mapped.get("coordinates"))
        coord_stats[pt.status] += 1
        source_id = result.mapped.get("source_id") or f"house_{idx:06d}"
        db.add(
            HousingRecord(
                source_id=str(source_id),
                name=result.mapped.get("name"),
                price=_float_or_none(result.mapped.get("price")),
                unit_price=_float_or_none(result.mapped.get("unit_price")),
                area=_float_or_none(result.mapped.get("area")),
                direction=result.mapped.get("direction"),
                room_type=result.mapped.get("room_type"),
                residence=result.mapped.get("residence"),
                building_type=result.mapped.get("building_type"),
                year=_int_or_none(result.mapped.get("year")),
                lng=pt.lng if pt.is_usable else None,
                lat=pt.lat if pt.is_usable else None,
                coord_status=pt.status,
                source_file=_rel_source(file_name),
                raw_json=result.raw_json,
            )
        )
        written += 1

    db.flush()
    return _file_report(
        file_name, "house_price", path, len(records), written,
        missing_counts, coord_stats, missing_optional_counts,
    )


# --------------------------------------------------------------------------- #
# 人口（总量 + 画像，按 grid_key 合并）
# --------------------------------------------------------------------------- #
def _read_json(file_name: str) -> list[dict]:
    path = settings.corpus_path / file_name
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _ingest_population(db: Session) -> list[dict[str, Any]]:
    total_name = "区域人口总量.json"
    profile_name = "区域人口画像.json"
    total_path = settings.corpus_path / total_name
    profile_path = settings.corpus_path / profile_name

    reports: list[dict[str, Any]] = []
    grids: dict[str, dict[str, Any]] = {}

    # ---- 总量 ----
    total_coord_stats: Counter[str] = Counter()
    total_missing: Counter[str] = Counter()
    total_records = _read_json(total_name)
    for raw in total_records:
        result = fms.map_record("population_total", raw)
        for mf in result.missing_fields:
            total_missing[mf] += 1
        bbox = geo_utils.parse_bbox(result.mapped.get("coordinates"))
        total_coord_stats[bbox.status] += 1
        if bbox.grid_key is None:
            continue
        grids.setdefault(bbox.grid_key, {})
        grids[bbox.grid_key].update(
            {
                "grid_id": bbox.grid_key,
                "residential": _int_or_none(result.mapped.get("residential")),
                "worker": _int_or_none(result.mapped.get("worker")),
                "center_lng": bbox.center_lng,
                "center_lat": bbox.center_lat,
                "bbox_geojson": bbox.bbox_geojson,
                "coord_status": bbox.status,
                "raw_total": result.raw_json,
            }
        )

    # ---- 画像 ----
    profile_coord_stats: Counter[str] = Counter()
    profile_records = _read_json(profile_name)
    matched_profile = 0
    for raw in profile_records:
        result = fms.map_record("population_profile", raw)
        bbox = geo_utils.parse_bbox(result.mapped.get("coordinates"))
        profile_coord_stats[bbox.status] += 1
        if bbox.grid_key is None:
            continue
        profile_fields = fms.extract_profile_fields(raw)
        node = grids.setdefault(bbox.grid_key, {"grid_id": bbox.grid_key})
        node["profile_json"] = json.dumps(profile_fields, ensure_ascii=False)
        node["raw_profile"] = result.raw_json
        if "raw_total" in node:
            matched_profile += 1

    # ---- 合并写入 ----
    written = 0
    unmatched = 0
    for grid_key, node in grids.items():
        has_total = "raw_total" in node
        has_profile = "profile_json" in node
        if not (has_total and has_profile):
            unmatched += 1
        raw_merged = json.dumps(
            {"total": node.get("raw_total"), "profile": node.get("raw_profile")},
            ensure_ascii=False,
        )
        db.add(
            PopulationProfile(
                grid_id=grid_key,
                residential=node.get("residential"),
                worker=node.get("worker"),
                center_lng=node.get("center_lng"),
                center_lat=node.get("center_lat"),
                bbox_geojson=node.get("bbox_geojson"),
                profile_json=node.get("profile_json"),
                coord_status=node.get("coord_status"),
                source_file=_rel_source(total_name),
                raw_json=raw_merged,
            )
        )
        written += 1
    db.flush()

    warnings: list[str] = []
    if unmatched:
        warnings.append(f"{unmatched} 个网格未能在总量/画像之间完整匹配")

    reports.append(
        _file_report(
            total_name, "population", total_path, len(total_records), written,
            total_missing, total_coord_stats, {},
            warnings=warnings, note="人口总量（与画像按 grid_key 合并写入 population_profiles）",
        )
    )
    reports.append(
        _file_report(
            profile_name, "population_profile_src", profile_path,
            len(profile_records), matched_profile,
            Counter(), profile_coord_stats, {},
            note="人口画像（合并入 population 网格，不单独建表）",
        )
    )
    return reports


# --------------------------------------------------------------------------- #
# 矢量数据样例及说明表（口径 spec，不参与 split）
# --------------------------------------------------------------------------- #
def _ingest_spec(db: Session, file_name: str) -> dict[str, Any]:
    import openpyxl

    path = settings.corpus_path / file_name
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets_summary = []
    for ws in wb.worksheets:
        sheets_summary.append(
            {"sheet": ws.title, "rows": ws.max_row, "cols": ws.max_column}
        )
    wb.close()

    db.add(
        DataFile(
            file_name=file_name,
            data_type="spec",
            source="corpus",
            record_count=len(sheets_summary),
            file_hash=_file_hash(path),
            is_sensitive=True,
            split_summary=json.dumps(
                {"role": "field_spec", "sheets": sheets_summary}, ensure_ascii=False
            ),
        )
    )
    db.flush()
    return {
        "file_name": file_name,
        "data_type": "spec",
        "source_file": _rel_source(file_name),
        "file_hash": _file_hash(path),
        "record_count_raw": len(sheets_summary),
        "record_count_written": 0,
        "skipped": 0,
        "missing_field_counts": {},
        "missing_optional_counts": {},
        "coordinate_stats": {},
        "warnings": [],
        "note": "字段口径说明表，仅作字段映射校验/规则参考，不参与 split",
        "sheets": sheets_summary,
    }


# --------------------------------------------------------------------------- #
# 报告辅助
# --------------------------------------------------------------------------- #
def _file_report(
    file_name: str,
    data_type: str,
    path: Path,
    raw_count: int,
    written: int,
    missing_counts: Counter,
    coord_stats: Counter,
    missing_optional_counts: Counter,
    warnings: list[str] | None = None,
    note: str | None = None,
) -> dict[str, Any]:
    report = {
        "file_name": file_name,
        "data_type": data_type,
        "source_file": _rel_source(file_name),
        "file_hash": _file_hash(path),
        "record_count_raw": raw_count,
        "record_count_written": written,
        "skipped": raw_count - written,
        "missing_field_counts": dict(missing_counts),
        "missing_optional_counts": dict(missing_optional_counts),
        "coordinate_stats": dict(coord_stats),
        "warnings": list(warnings or []),
    }
    if note:
        report["note"] = note
    return report


def _write_datafile(db: Session, file_report: dict[str, Any]) -> None:
    """为一个结构化数据文件写入 DataFile 记录。"""
    db.add(
        DataFile(
            file_name=file_report["file_name"],
            data_type=file_report["data_type"],
            source="corpus",
            record_count=file_report["record_count_written"],
            file_hash=file_report["file_hash"],
            is_sensitive=True,
            split_summary=None,  # 由 split_manager 后续回写
        )
    )


# --------------------------------------------------------------------------- #
# 入口
# --------------------------------------------------------------------------- #
def run_ingestion(db: Session) -> dict[str, Any]:
    """执行完整导入流程，返回质量报告摘要（脱敏）。"""
    init_db()
    corpus = settings.corpus_path
    if not corpus.exists():
        raise FileNotFoundError(f"训练语料目录不存在: {corpus.name}")

    _clear_tables(db)

    file_reports: list[dict[str, Any]] = []
    records_written: dict[str, int] = {}

    # 点数据
    for file_name, data_type in (
        ("POI兴趣点分布数据.json", "poi"),
        ("产业布局数据.json", "industry"),
    ):
        if (corpus / file_name).exists():
            rep = _ingest_point_file(db, file_name, data_type)
            file_reports.append(rep)
            _write_datafile(db, rep)
            records_written[data_type] = rep["record_count_written"]

    # 房价
    if (corpus / "房价历史交易数据.json").exists():
        rep = _ingest_house_file(db, "房价历史交易数据.json")
        file_reports.append(rep)
        _write_datafile(db, rep)
        records_written["house_price"] = rep["record_count_written"]

    # 人口（合并）
    if (corpus / "区域人口总量.json").exists() and (corpus / "区域人口画像.json").exists():
        pop_reports = _ingest_population(db)
        file_reports.extend(pop_reports)
        # 仅为总量文件写一条 population DataFile（画像合并入网格）
        _write_datafile(db, pop_reports[0])
        # 画像源文件也登记一条 DataFile（记录来源，但不单独建表/切分）
        db.add(
            DataFile(
                file_name=pop_reports[1]["file_name"],
                data_type="population_profile_src",
                source="corpus",
                record_count=pop_reports[1]["record_count_written"],
                file_hash=pop_reports[1]["file_hash"],
                is_sensitive=True,
                split_summary=json.dumps(
                    {"role": "merged_into", "target": "population"}, ensure_ascii=False
                ),
            )
        )
        records_written["population"] = pop_reports[0]["record_count_written"]

    # 口径说明表
    if (corpus / "矢量数据样例及说明表.xlsx").exists():
        spec_rep = _ingest_spec(db, "矢量数据样例及说明表.xlsx")
        file_reports.append(spec_rep)

    db.commit()

    report = {
        "created_at": _utcnow_iso(),
        "mode": settings.app_mode,
        "corpus_dir_name": corpus.name,
        "files": file_reports,
        "totals": {
            "files_processed": len(file_reports),
            "records_written": records_written,
        },
        "notes": [
            "本报告仅含统计量，不含任何语料原文。",
            "test split 仅在 split_manager 生成并冻结，本阶段不用于训练/分析。",
        ],
    }

    settings.processed_dir.mkdir(parents=True, exist_ok=True)
    with settings.quality_report_path.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    logger.info(
        "ingestion done: files=%s records=%s",
        report["totals"]["files_processed"],
        records_written,
    )
    return report


def get_status(db: Session) -> dict[str, Any]:
    """返回导入状态：各表计数 + 质量报告是否存在。"""
    counts = {
        "poi": db.query(PoiPoint).count(),
        "industry": db.query(IndustryPoint).count(),
        "house_price": db.query(HousingRecord).count(),
        "population": db.query(PopulationProfile).count(),
    }
    data_files = db.query(DataFile).filter(DataFile.source == "corpus").count()
    qr = settings.quality_report_path
    return {
        "ingested": any(counts.values()),
        "table_counts": counts,
        "data_files": data_files,
        "quality_report_exists": qr.exists(),
        "quality_report_path": str(qr.relative_to(settings.data_dir.parent)),
        "mode": settings.app_mode,
    }


def get_quality_report() -> dict[str, Any] | None:
    qr = settings.quality_report_path
    if not qr.exists():
        return None
    with qr.open("r", encoding="utf-8") as f:
        return json.load(f)
