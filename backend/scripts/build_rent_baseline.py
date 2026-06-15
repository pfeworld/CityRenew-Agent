"""一次性预处理：科研授权链家租赁挂牌 → 全市租金参考基线（元/㎡/月）。

红线：
- 仅本地处理，输出统计量（中位/分位/样本量），不落原文、不外发、不入公开仓库。
- 该租赁数据为"去除小区信息版"，无法稳定定位到行政区，按红线只给"全市参考"口径并标注。
- 产物 rent_baseline.json 落在 backend/data/external/authorized_property/processed/（已 gitignore）。

运行：.venv/bin/python -m scripts.build_rent_baseline
"""
from __future__ import annotations

import json
import re
from datetime import datetime, timezone

import numpy as np

from app.config import settings

_ROOT = settings.data_dir.parent.parent  # backend/data -> backend -> 项目根
_CANDIDATES = [
    _ROOT / "科研语料" / "上海链家租赁挂牌案例(去除社区信息版).xlsx",
    _ROOT / "科研语料" / "上海链家租赁挂牌案例(去除社区信息版)-7.23.xlsx",
]
_OUT = (settings.data_dir / "external" / "authorized_property" / "processed"
        / "rent_baseline.json")

# 合理租金单价带（元/㎡/月）：剔除明显异常/口径不一致记录，不臆测原值。
RENT_UNIT_MIN = 20.0
RENT_UNIT_MAX = 800.0


def _num(text) -> float | None:
    if text is None:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)", str(text))
    return float(m.group(1)) if m else None


def _rent_total(text) -> float | None:
    """从『出租价格及支付方式』提取月租金额（元/月）。"""
    if text is None:
        return None
    m = re.search(r"(\d{3,6})\s*元", str(text))
    if m:
        return float(m.group(1))
    return _num(text)


def main() -> None:
    import openpyxl

    src = next((p for p in _CANDIDATES if p.exists()), None)
    if src is None:
        print("未找到链家租赁文件：", [str(p) for p in _CANDIDATES])
        return

    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    price_col = idx.get("出租价格及支付方式")
    area_col = idx.get("面积")
    if price_col is None or area_col is None:
        print("缺少必要列：", header)
        wb.close()
        return

    units: list[float] = []
    rents: list[float] = []
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n += 1
        rent = _rent_total(row[price_col])
        area = _num(row[area_col])
        if not rent or not area or area <= 0:
            continue
        u = rent / area
        if RENT_UNIT_MIN <= u <= RENT_UNIT_MAX:
            units.append(round(u, 2))
            rents.append(rent)
    wb.close()

    if len(units) < 100:
        print(f"有效租金样本过少（{len(units)}），不生成基线。")
        return

    arr = np.array(units, dtype=float)
    obj = {
        "scope": "citywide",
        "scope_note": "链家租赁挂牌为去除小区信息版，无法稳定定位到行政区，仅作全市租金参考。",
        "median_rent_unit": round(float(np.median(arr)), 2),
        "p25_rent_unit": round(float(np.percentile(arr, 25)), 2),
        "p75_rent_unit": round(float(np.percentile(arr, 75)), 2),
        "median_rent_total": round(float(np.median(np.array(rents))), 0),
        "sample_count": int(len(units)),
        "unit": "元/㎡/月",
        "source": "科研授权脱敏链家租赁挂牌案例（去除小区信息版）",
        "filter_rule": f"仅保留 {RENT_UNIT_MIN:.0f}~{RENT_UNIT_MAX:.0f} 元/㎡/月 的口径一致样本",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    _OUT.parent.mkdir(parents=True, exist_ok=True)
    _OUT.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"已写入 {_OUT}")
    print(f"全市租金参考：中位 {obj['median_rent_unit']} 元/㎡/月 "
          f"(p25 {obj['p25_rent_unit']} / p75 {obj['p75_rent_unit']})，样本 {obj['sample_count']} 条")


if __name__ == "__main__":
    main()
